"""
ls_to_xlsx.py
Converts a Label Studio annotation JSON file into a two-tab Excel workbook.

Usage:
    python ls_to_xlsx.py <input.json>

Output:
    <input>.xlsx  (same base name, same directory as input file)
"""

import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill("solid", start_color="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
CELL_FONT   = Font(name="Arial", size=10)
WRAP        = Alignment(wrap_text=True, vertical="top")
TOP         = Alignment(vertical="top")


def style_header(ws, columns):
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font    = HEADER_FONT
        cell.fill    = HEADER_FILL
        cell.alignment = TOP


def style_data_cell(cell, wrap=False):
    cell.font      = CELL_FONT
    cell.alignment = WRAP if wrap else TOP


def set_column_widths(ws, widths):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def parse_annotation(annotation):
    """
    Return (analytic_rows, holistic_row) for a single annotation object.

    analytic_rows : list of dicts, one per error span
    holistic_row  : dict with holistic fields (may be empty if none marked)
    """
    completed_by = annotation["completed_by"]
    result       = annotation.get("result", [])

    # --- group result items by span id ---
    spans = defaultdict(dict)   # span_id -> {from_name: value}
    holistic = {}

    HOLISTIC_KEYS = {
        "overall_correspondence", "correspondence_comments",
        "overall_readability", "readability_comments", "document_issues"
    }

    for item in result:
        from_name = item.get("from_name", "")
        value     = item.get("value", {})
        span_id   = item.get("id")

        if from_name in HOLISTIC_KEYS:
            if "rating" in value:
                holistic[from_name] = value["rating"]
            elif "text" in value:
                holistic[from_name] = value["text"][0] if value["text"] else ""
            continue

        # Analytic items — keyed by span id
        if from_name == "label":
            spans[span_id]["start"] = value.get("start")
            spans[span_id]["end"]   = value.get("end")
            spans[span_id]["text"]  = value.get("text", "")
            spans[span_id]["label"] = value["labels"][0] if value.get("labels") else ""

        elif from_name == "subcategories":
            choices = value.get("choices", [])
            spans[span_id]["subcategory"] = choices[0] if choices else ""

        elif from_name == "impact":
            choices = value.get("choices", [])
            spans[span_id]["impact"] = choices[0] if choices else ""

        elif from_name == "comments":
            texts = value.get("text", [])
            spans[span_id]["comment"] = texts[0] if texts else ""

    # Build analytic rows (only spans that have a label entry)
    analytic_rows = []
    for span_id, fields in spans.items():
        if "label" not in fields:
            continue
        analytic_rows.append({
            "completed_by": completed_by,
            "start":        fields.get("start", ""),
            "end":          fields.get("end", ""),
            "text":         fields.get("text", ""),
            "label":        fields.get("label", ""),
            "subcategory":  fields.get("subcategory", ""),
            "impact":       fields.get("impact", ""),
            "comment":      fields.get("comment", ""),
        })

    holistic_row = {
        "completed_by":             completed_by,
        "overall_correspondence":   holistic.get("overall_correspondence", ""),
        "correspondence_comments":  holistic.get("correspondence_comments", ""),
        "overall_readability":      holistic.get("overall_readability", ""),
        "readability_comments":     holistic.get("readability_comments", ""),
        "document_issues":          holistic.get("document_issues", ""),
    }

    return analytic_rows, holistic_row


def build_workbook(tasks):
    wb = Workbook()

    # --- Analytic sheet ---
    ws_a = wb.active
    ws_a.title = "Analytic"

    analytic_cols = [
        "completed_by", "start", "end", "text",
        "label", "subcategory", "impact", "comment"
    ]
    style_header(ws_a, analytic_cols)

    # --- Holistic sheet ---
    ws_h = wb.create_sheet("Holistic")

    holistic_cols = [
        "completed_by",
        "overall_correspondence", "correspondence_comments",
        "overall_readability", "readability_comments",
        "document_issues"
    ]
    style_header(ws_h, holistic_cols)

    for task in tasks:
        for annotation in task.get("annotations", []):
            analytic_rows, holistic_row = parse_annotation(annotation)

            for row in analytic_rows:
                ws_a.append([row[c] for c in analytic_cols])
                r = ws_a.max_row
                for col_idx, col in enumerate(analytic_cols, start=1):
                    style_data_cell(ws_a.cell(r, col_idx), wrap=(col in ("text", "comment")))

            ws_h.append([holistic_row[c] for c in holistic_cols])
            r = ws_h.max_row
            for col_idx, col in enumerate(holistic_cols, start=1):
                style_data_cell(ws_h.cell(r, col_idx),
                                wrap=(col in ("correspondence_comments",
                                              "readability_comments",
                                              "document_issues")))

    # Column widths
    set_column_widths(ws_a, {
        "A": 14,  # completed_by
        "B": 8,   # start
        "C": 8,   # end
        "D": 50,  # text
        "E": 22,  # label
        "F": 22,  # subcategory
        "G": 12,  # impact
        "H": 45,  # comment
    })
    set_column_widths(ws_h, {
        "A": 14,  # completed_by
        "B": 24,  # overall_correspondence
        "C": 45,  # correspondence_comments
        "D": 20,  # overall_readability
        "E": 45,  # readability_comments
        "F": 50,  # document_issues
    })

    # Freeze header rows
    ws_a.freeze_panes = "A2"
    ws_h.freeze_panes = "A2"

    return wb


def main():
    if len(sys.argv) != 2:
        print("Usage: python ls_to_xlsx.py <input.json>")
        sys.exit(1)

    input_path  = Path(sys.argv[1])
    output_path = input_path.with_suffix(".xlsx")

    with open(input_path, encoding="utf-8") as f:
        tasks = json.load(f)

    wb = build_workbook(tasks)
    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
